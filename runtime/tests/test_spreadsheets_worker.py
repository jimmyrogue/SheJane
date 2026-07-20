from __future__ import annotations

import asyncio
import hashlib
import json
import os
import subprocess
import sys
import uuid
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from shejane_runtime.plugins.executor import ManagedWorkerActionExecutor
from shejane_runtime.plugins.macos_vm import load_macos_vm_resources
from shejane_runtime.plugins.platforms import (
    current_managed_worker_execution_platform,
    current_managed_worker_platform,
)
from shejane_runtime.plugins.runtime_assets import RuntimeAssetHandle

REPO_ROOT = Path(__file__).resolve().parents[2]
WORKER = (
    REPO_ROOT
    / "runtime"
    / "plugins"
    / "office"
    / "spreadsheets"
    / "worker"
    / "spreadsheets_worker.py"
)
GOLDEN_ROOT = REPO_ROOT / "runtime" / "plugins" / "office" / "spreadsheets" / "golden"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _worker_command() -> tuple[str, ...]:
    frozen = os.environ.get("SHEJANE_TEST_SPREADSHEETS_WORKER")
    return (frozen,) if frozen else (sys.executable, str(WORKER))


def _invocation(
    action_id: str,
    arguments: dict[str, object],
    *,
    source: Path | None = None,
) -> dict[str, object]:
    invocation: dict[str, object] = {
        "schema_version": 1,
        "invocation_id": str(uuid.uuid4()),
        "operation_id": f"run_01:{action_id}:001",
        "action": {
            "plugin_id": "org.shejane.spreadsheets",
            "plugin_digest": "sha256:" + "e" * 64,
            "action_id": action_id,
        },
        "arguments": arguments,
        "inputs": [],
        "grants": {"capabilities": []},
        "limits": {"timeout_ms": 10_000, "memory_mb": 512, "output_mb": 64},
    }
    if source is not None:
        data = source.read_bytes()
        invocation["inputs"] = [
            {
                "id": "workbook",
                "path": f"/input/{source.name}",
                "media_type": XLSX_MIME,
                "size_bytes": len(data),
                "sha256": hashlib.sha256(data).hexdigest(),
            }
        ]
    return invocation


async def _invoke(
    tmp_path: Path,
    action_id: str,
    arguments: dict[str, object],
    *,
    source: Path | None = None,
    runtime_assets: tuple[RuntimeAssetHandle, ...] = (),
) -> tuple[dict[str, object], Path]:
    roots = tmp_path / str(uuid.uuid4())
    input_root = roots / "input"
    output_root = roots / "output"
    input_root.mkdir(parents=True)
    output_root.mkdir()
    materialized = None
    if source is not None:
        materialized = input_root / source.name
        materialized.write_bytes(source.read_bytes())
    vm_manifest = os.environ.get("SHEJANE_TEST_MACOS_VM_ASSETS")
    command = _worker_command()
    result = await ManagedWorkerActionExecutor(
        command,
        runtime_assets=runtime_assets,
        vm_resources=(
            load_macos_vm_resources(Path(vm_manifest).resolve(strict=True)) if vm_manifest else None
        ),
        package_root=Path(command[0]).resolve(strict=True).parent if vm_manifest else None,
    ).invoke(
        _invocation(action_id, arguments, source=materialized),
        input_root=input_root,
        output_root=output_root,
    )
    return result, output_root


@pytest.mark.asyncio
async def test_spreadsheets_worker_create_read_edit_and_bound_ranges(tmp_path: Path) -> None:
    create_arguments = {
        "filename": "forecast.xlsx",
        "worksheets": [
            {
                "name": "Summary",
                "cells": [
                    {"cell": "A1", "value": "Month", "style": {"bold": True}},
                    {"cell": "B1", "value": "Revenue", "style": {"bold": True}},
                    {
                        "cell": "A2",
                        "value": "2026-07-16T08:00:00+08:00",
                        "value_type": "datetime",
                        "number_format": "yyyy-mm-dd hh:mm",
                    },
                    {"cell": "B2", "value": 0.1, "number_format": "0.000"},
                    {"cell": "B3", "value": 0.2, "number_format": "0.000"},
                    {"cell": "B4", "formula": "=SUM(B2:B3)", "number_format": "0.000"},
                ],
                "merges": ["A6:B6"],
                "charts": [
                    {
                        "type": "bar",
                        "title": "Revenue",
                        "data_range": "B1:B3",
                        "categories_range": "A1:A3",
                        "anchor": "D2",
                    }
                ],
            }
        ],
    }
    created, create_output = await _invoke(tmp_path, "spreadsheet.create", create_arguments)
    replay, replay_output = await _invoke(tmp_path, "spreadsheet.create", create_arguments)
    created_path = create_output / "forecast.xlsx"
    assert created["status"] == "succeeded"
    assert replay["status"] == "succeeded"
    assert created_path.read_bytes() == (replay_output / "forecast.xlsx").read_bytes()

    original_digest = hashlib.sha256(created_path.read_bytes()).hexdigest()
    read, _ = await _invoke(
        tmp_path,
        "spreadsheet.read",
        {"input_id": "workbook", "sheet": "Summary", "range": "A1:B100", "max_cells": 8},
        source=created_path,
    )
    assert read["status"] == "succeeded"
    output = read["output"]
    assert output["sheets"] == [{"name": "Summary", "max_row": 6, "max_column": 2}]
    assert output["requested_range"] == "A1:B100"
    assert output["range"] == "A1:B4"
    assert output["truncated"] is True
    assert output["formulas"][3][1] == "=SUM(B2:B3)"
    assert output["values"][3][1] is None
    assert output["types"][1][0] == "datetime"
    assert output["values"][1][0] == "2026-07-16T00:00:00Z"
    assert output["number_formats"][1][1] == "0.000"

    edited, edit_output = await _invoke(
        tmp_path,
        "spreadsheet.edit",
        {
            "input_id": "workbook",
            "output_filename": "forecast-edited.xlsx",
            "operations": [
                {
                    "type": "set_cells",
                    "sheet": "Summary",
                    "range": "A3:B3",
                    "values": [["Aug", 42.25]],
                },
                {"type": "set_formula", "sheet": "Summary", "cell": "B4", "formula": "=SUM(B2:B3)"},
                {
                    "type": "format_range",
                    "sheet": "Summary",
                    "range": "A1:B1",
                    "style": {"bold": True, "fill_color": "D9EAD3", "horizontal": "center"},
                },
                {"type": "append_rows", "sheet": "Summary", "rows": [["Sep", 50.5]]},
            ],
        },
        source=created_path,
    )
    assert edited["status"] == "succeeded"
    assert hashlib.sha256(created_path.read_bytes()).hexdigest() == original_digest
    workbook = load_workbook(edit_output / "forecast-edited.xlsx", data_only=False)
    try:
        sheet = workbook["Summary"]
        assert sheet["B3"].value == 42.25
        assert sheet["B4"].value == "=SUM(B2:B3)"
        assert sheet["A1"].fill.fgColor.rgb == "00D9EAD3"
        assert sheet["A7"].value == "Sep"
        assert len(sheet._charts) == 1
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_spreadsheets_worker_returns_stable_input_errors(tmp_path: Path) -> None:
    for name, data, code in (
        ("broken.xlsx", b"not a zip", "spreadsheet_corrupt"),
        ("encrypted.xlsx", bytes.fromhex("d0cf11e0a1b11ae1"), "spreadsheet_encrypted_or_legacy"),
    ):
        source = tmp_path / name
        source.write_bytes(data)
        result, _ = await _invoke(
            tmp_path,
            "spreadsheet.read",
            {"input_id": "workbook"},
            source=source,
        )
        assert result["status"] == "failed"
        assert result["error"]["code"] == code  # type: ignore[index]


@pytest.mark.asyncio
async def test_spreadsheets_worker_recalculates_and_renders_with_exact_asset(
    tmp_path: Path,
) -> None:
    source = tmp_path / "render.xlsx"
    created, created_output = await _invoke(
        tmp_path,
        "spreadsheet.create",
        {
            "filename": source.name,
            "worksheets": [
                {
                    "name": "Sheet1",
                    "cells": [
                        {"cell": "A1", "value": 1},
                        {"cell": "A2", "value": 2},
                        {"cell": "A3", "formula": "=SUM(A1:A2)"},
                    ],
                }
            ],
        },
    )
    assert created["status"] == "succeeded"
    source.write_bytes((created_output / source.name).read_bytes())

    asset_root = tmp_path / "office-runtime"
    asset_root.mkdir()
    soffice = asset_root / "soffice"
    mutool = asset_root / "mutool"
    soffice.write_text(
        f"#!{sys.executable}\n"
        "import pathlib, shutil, sys\n"
        "args = sys.argv[1:]\n"
        "out = pathlib.Path(args[args.index('--outdir') + 1])\n"
        "source = pathlib.Path(args[-1])\n"
        "if any(arg.startswith('xlsx') for arg in args):\n"
        "    shutil.copyfile(source, out / source.name)\n"
        "else:\n"
        "    (out / (source.stem + '.pdf')).write_bytes(b'%PDF-1.7\\n%%EOF\\n')\n",
        encoding="utf-8",
    )
    mutool.write_text(
        f"#!{sys.executable}\n"
        "import pathlib, sys\n"
        "args = sys.argv[1:]\n"
        "if args[0] == 'info':\n"
        "    print('Pages: 1')\n"
        "else:\n"
        "    pathlib.Path(args[args.index('-o') + 1] % 1).write_bytes(b'PNG')\n",
        encoding="utf-8",
    )
    soffice.chmod(0o500)
    mutool.chmod(0o500)
    (asset_root / "office-runtime.json").write_text(
        json.dumps({"soffice": "soffice", "mutool": "mutool"}),
        encoding="utf-8",
    )
    sbom = asset_root / "sbom.spdx.json"
    sbom.write_text("{}", encoding="utf-8")
    asset = RuntimeAssetHandle(
        asset_id="org.libreoffice.runtime",
        version="25.8.7",
        platform="darwin/arm64",
        digest="sha256:" + "a" * 64,
        root=asset_root,
        payload=asset_root,
        license="MPL-2.0",
        source_url="https://www.libreoffice.org/",
        sbom=sbom,
    )

    rendered, output_root = await _invoke(
        tmp_path,
        "spreadsheet.render",
        {"input_id": "workbook", "include_png": True, "max_pages": 1, "dpi": 144},
        source=source,
        runtime_assets=(asset,),
    )
    assert rendered["status"] == "succeeded", rendered
    assert rendered["output"] == {
        "page_count": 1,
        "rendered_pages": 1,
        "recalculated_name": "render.recalculated.xlsx",
        "pdf_name": "render.recalculated.pdf",
        "png_names": ["render.recalculated.page-0001.png"],
        "warnings": [],
    }
    assert {path.name for path in output_root.iterdir()} == {
        "render.recalculated.xlsx",
        "render.recalculated.pdf",
        "render.recalculated.page-0001.png",
    }


@pytest.mark.asyncio
async def test_spreadsheets_worker_real_runtime_recalculation_and_golden(tmp_path: Path) -> None:
    configured = os.environ.get("SHEJANE_TEST_OFFICE_RUNTIME")
    if not configured:
        pytest.skip("real Office runtime asset not configured")
    asset_root = Path(configured).resolve(strict=True)
    target = (
        current_managed_worker_execution_platform()
        if os.environ.get("SHEJANE_TEST_MACOS_VM_ASSETS")
        else current_managed_worker_platform()
    )
    assert target is not None
    asset = RuntimeAssetHandle(
        asset_id="org.libreoffice.runtime",
        version="25.8.7",
        platform=target,
        digest="sha256:" + "a" * 64,
        root=asset_root,
        payload=asset_root / "payload",
        license="MPL-2.0 AND AGPL-3.0-only AND OFL-1.1",
        source_url="https://www.libreoffice.org/",
        sbom=asset_root / "payload" / "office-runtime.json",
    )
    create_arguments = {
        "filename": "golden.xlsx",
        "worksheets": [
            {
                "name": "计算表",
                "cells": [
                    {
                        "cell": "A1",
                        "value": "项目",
                        "style": {"bold": True, "fill_color": "D9EAD3"},
                    },
                    {"cell": "B1", "value": "值", "style": {"bold": True, "fill_color": "D9EAD3"}},
                    {"cell": "A2", "value": "Decimal A"},
                    {"cell": "B2", "value": 0.1, "number_format": "0.000"},
                    {"cell": "A3", "value": "Decimal B"},
                    {"cell": "B3", "value": 0.2, "number_format": "0.000"},
                    {"cell": "A4", "value": "Formula sum"},
                    {"cell": "B4", "formula": "=SUM(B2:B3)", "number_format": "0.000"},
                    {"cell": "A5", "value": "UTC time"},
                    {
                        "cell": "B5",
                        "value": "2026-07-16T08:00:00+08:00",
                        "value_type": "datetime",
                        "number_format": "yyyy-mm-dd hh:mm",
                    },
                    {"cell": "A6", "value": "Locale number"},
                    {"cell": "B6", "value": 1234.5, "number_format": "[$-409]#,##0.00"},
                    {"cell": "A8", "value": "石间 · 日本語 · 한국어"},
                ],
                "merges": ["A8:B8"],
                "column_widths": {"A": 20, "B": 22},
                "charts": [
                    {
                        "type": "bar",
                        "title": "Decimal values",
                        "data_range": "B1:B3",
                        "categories_range": "A2:A3",
                        "anchor": "D2",
                    }
                ],
            }
        ],
    }
    created, created_output = await _invoke(tmp_path, "spreadsheet.create", create_arguments)
    assert created["status"] == "succeeded"
    source = created_output / "golden.xlsx"
    source_digest = hashlib.sha256(source.read_bytes()).hexdigest()

    rendered, output = await _invoke(
        tmp_path,
        "spreadsheet.render",
        {"input_id": "workbook", "include_png": True, "max_pages": 4, "dpi": 144},
        source=source,
        runtime_assets=(asset,),
    )
    assert rendered["status"] == "succeeded", rendered
    assert hashlib.sha256(source.read_bytes()).hexdigest() == source_digest
    recalculated = output / "golden.recalculated.xlsx"
    formulas = load_workbook(recalculated, data_only=False, read_only=True)
    values = load_workbook(recalculated, data_only=True, read_only=True)
    try:
        assert formulas["计算表"]["B4"].value == "=SUM(B2:B3)"
        assert values["计算表"]["B4"].value == pytest.approx(0.3)
        assert values["计算表"]["B5"].value.isoformat() == "2026-07-16T00:00:00"
        assert formulas["计算表"]["B6"].number_format == "[$-409]#,##0.00"
    finally:
        formulas.close()
        values.close()

    if not os.environ.get("SHEJANE_TEST_MACOS_VM_ASSETS"):
        config = json.loads((asset.payload / "office-runtime.json").read_text(encoding="utf-8"))
        mutool = asset.payload / config["mutool"]
        extracted = (
            await asyncio.to_thread(
                subprocess.run,
                [
                    str(mutool),
                    "draw",
                    "-F",
                    "txt",
                    "-o",
                    "-",
                    str(output / "golden.recalculated.pdf"),
                ],
                check=True,
                capture_output=True,
                text=True,
            )
        ).stdout
        assert "Formula sum" in extracted
        assert "石间" in extracted

    expected = json.loads((GOLDEN_ROOT / f"{target.replace('/', '-')}.json").read_text())
    assert rendered["output"]["page_count"] == expected["page_count"]  # type: ignore[index]
    actual = []
    for page in range(1, rendered["output"]["rendered_pages"] + 1):  # type: ignore[index,operator]
        with Image.open(output / f"golden.recalculated.page-{page:04d}.png") as image:
            gray = image.convert("L")
            actual.append(
                {
                    "size": list(gray.size),
                    "ink_bbox": list(_ink_bbox(gray)),
                    "dhash": _dhash(gray),
                }
            )
    assert actual == expected["pages"]


def _ink_bbox(image: Image.Image) -> tuple[int, int, int, int]:
    return image.point(lambda value: 255 - value).getbbox() or (0, 0, 0, 0)


def _dhash(image: Image.Image) -> str:
    sampled = image.resize((9, 8), Image.Resampling.LANCZOS)
    pixels = list(sampled.get_flattened_data())
    value = 0
    for row in range(8):
        for column in range(8):
            value = (value << 1) | (pixels[row * 9 + column] > pixels[row * 9 + column + 1])
    return f"{value:016x}"
