#!/usr/bin/env python3
"""One-shot Spreadsheets Managed Worker for SheJane Plugin Action Protocol v1."""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import zipfile
from datetime import date, datetime, time as datetime_time, timezone
from pathlib import Path, PurePosixPath
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import range_boundaries

PLUGIN_ID = "org.shejane.spreadsheets"
RUNTIME_ASSET_ID = "org.libreoffice.runtime"
ACTION_IDS = {
    "spreadsheet.read",
    "spreadsheet.create",
    "spreadsheet.edit",
    "spreadsheet.render",
}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
MAX_CELLS = 10_000
MARKDOWN_LIMIT = 60_000
_CELL_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}$")
_RANGE_RE = re.compile(
    r"^[A-Za-z]{1,3}[1-9][0-9]{0,6}(?::[A-Za-z]{1,3}[1-9][0-9]{0,6})?$"
)


class SpreadsheetActionError(RuntimeError):
    def __init__(self, code: str, message: str, *, retryable: bool = False) -> None:
        super().__init__(message)
        self.code = code
        self.retryable = retryable


def _reply(request_id: int, result: object) -> None:
    print(
        json.dumps(
            {"jsonrpc": "2.0", "id": request_id, "result": result},
            ensure_ascii=False,
            separators=(",", ":"),
        ),
        flush=True,
    )


def _request(expected_method: str) -> dict[str, Any]:
    try:
        payload = json.loads(sys.stdin.readline())
    except (EOFError, json.JSONDecodeError) as exc:
        raise SpreadsheetActionError("protocol_violation", "invalid protocol frame") from exc
    if (
        not isinstance(payload, dict)
        or payload.get("jsonrpc") != "2.0"
        or payload.get("method") != expected_method
        or not isinstance(payload.get("params"), dict)
    ):
        raise SpreadsheetActionError("protocol_violation", f"expected {expected_method}")
    return payload


def _output_root() -> Path:
    try:
        return Path(os.environ["SHEJANE_PLUGIN_OUTPUT_ROOT"]).resolve(strict=True)
    except (KeyError, OSError) as exc:
        raise SpreadsheetActionError(
            "invalid_invocation", "output staging is unavailable"
        ) from exc


def _materialized_input(invocation: dict[str, Any], input_id: str) -> Path:
    references = invocation.get("inputs")
    if not isinstance(references, list):
        raise SpreadsheetActionError("invalid_invocation", "workbook input is unavailable")
    reference = next(
        (item for item in references if isinstance(item, dict) and item.get("id") == input_id),
        None,
    )
    if reference is None or reference.get("media_type") != XLSX_MIME:
        raise SpreadsheetActionError("invalid_invocation", "an XLSX input is required")
    try:
        relative = PurePosixPath(str(reference["path"])).relative_to("/input")
    except (KeyError, ValueError) as exc:
        raise SpreadsheetActionError(
            "invalid_invocation", "workbook input path is invalid"
        ) from exc
    if not relative.parts or any(part in {"", ".", ".."} for part in relative.parts):
        raise SpreadsheetActionError("invalid_invocation", "workbook input path is invalid")
    try:
        root = Path(os.environ["SHEJANE_PLUGIN_INPUT_ROOT"]).resolve(strict=True)
        source = root.joinpath(*relative.parts)
        if source.is_symlink() or not source.is_file():
            raise OSError
        source = source.resolve(strict=True)
        source.relative_to(root)
        data = source.read_bytes()
    except (KeyError, OSError, ValueError) as exc:
        raise SpreadsheetActionError("invalid_invocation", "workbook input is unavailable") from exc
    if len(data) != reference.get("size_bytes") or hashlib.sha256(
        data
    ).hexdigest() != reference.get("sha256"):
        raise SpreadsheetActionError("invalid_invocation", "workbook input digest changed")
    return source


def _open_workbook(path: Path, *, data_only: bool = False, read_only: bool = False):
    try:
        prefix = path.read_bytes()[:8]
    except OSError as exc:
        raise SpreadsheetActionError("invalid_invocation", "workbook input is unavailable") from exc
    if prefix.startswith(bytes.fromhex("d0cf11e0")):
        raise SpreadsheetActionError(
            "spreadsheet_encrypted_or_legacy",
            "encrypted or legacy binary Excel workbooks are unsupported",
        )
    if not prefix.startswith(b"PK"):
        raise SpreadsheetActionError("spreadsheet_corrupt", "XLSX package is corrupt")
    try:
        return load_workbook(path, data_only=data_only, read_only=read_only)
    except Exception as exc:
        raise SpreadsheetActionError("spreadsheet_corrupt", "XLSX package is corrupt") from exc


def _safe_output_name(name: str) -> str:
    candidate = PurePosixPath(name)
    if (
        not name.lower().endswith(".xlsx")
        or candidate.name != name
        or name in {".", ".."}
        or "\\" in name
    ):
        raise SpreadsheetActionError("invalid_invocation", "output filename must be a basename.xlsx")
    return name


def _range(value: str) -> tuple[int, int, int, int]:
    if not _RANGE_RE.fullmatch(value):
        raise SpreadsheetActionError("invalid_invocation", "cell range must use A1 notation")
    try:
        min_col, min_row, max_col, max_row = range_boundaries(value.upper())
    except ValueError as exc:
        raise SpreadsheetActionError("invalid_invocation", "cell range is invalid") from exc
    if max_col > 16_384 or max_row > 1_048_576:
        raise SpreadsheetActionError("invalid_invocation", "cell range exceeds XLSX limits")
    return min_row, min_col, max_row, max_col


def _limited_range(value: str) -> tuple[int, int, int, int]:
    parsed = _range(value)
    min_row, min_col, max_row, max_col = parsed
    if (max_row - min_row + 1) * (max_col - min_col + 1) > MAX_CELLS:
        raise SpreadsheetActionError("invalid_invocation", "cell range exceeds the action limit")
    return parsed


def _cell(value: str) -> str:
    if not _CELL_RE.fullmatch(value):
        raise SpreadsheetActionError("invalid_invocation", "cell must use A1 notation")
    _range(value)
    return value.upper()


def _column(value: str) -> str:
    candidate = value.upper()
    try:
        index = column_index_from_string(candidate)
    except ValueError as exc:
        raise SpreadsheetActionError("invalid_invocation", "column must use A1 letters") from exc
    if index > 16_384:
        raise SpreadsheetActionError("invalid_invocation", "column exceeds XLSX limits")
    return candidate


def _sheet(workbook, name: str | None):
    if name is None:
        return workbook.active
    if name not in workbook.sheetnames:
        raise SpreadsheetActionError("spreadsheet_sheet_not_found", "worksheet was not found")
    return workbook[name]


def _bounded_range(
    requested: str,
    max_cells: int,
) -> tuple[int, int, int, int, str, bool]:
    min_row, min_col, max_row, max_col = _range(requested)
    width = max_col - min_col + 1
    height = max_row - min_row + 1
    bounded_width = min(width, max_cells)
    bounded_height = min(height, max(1, max_cells // bounded_width))
    bounded_max_col = min_col + bounded_width - 1
    bounded_max_row = min_row + bounded_height - 1
    actual = (
        f"{get_column_letter(min_col)}{min_row}:"
        f"{get_column_letter(bounded_max_col)}{bounded_max_row}"
    )
    return min_row, min_col, bounded_max_row, bounded_max_col, actual, width * height > max_cells


def _json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z")
    if isinstance(value, (date, datetime_time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _value_type(value: Any, *, formula: bool, error: bool) -> str:
    if formula:
        return "formula"
    if error:
        return "error"
    if value is None:
        return "empty"
    if isinstance(value, datetime):
        return "datetime"
    if isinstance(value, date):
        return "date"
    if isinstance(value, datetime_time):
        return "time"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    return "string"


def _markdown(values: list[list[Any]], formulas: list[list[str | None]]) -> tuple[str, bool]:
    rows: list[str] = []
    for row_index, (value_row, formula_row) in enumerate(zip(values, formulas, strict=True)):
        rendered = []
        for value, formula in zip(value_row, formula_row, strict=True):
            text = str(formula if formula is not None and value is None else value or "")
            rendered.append(text.replace("|", "\\|").replace("\n", " "))
        rows.append("| " + " | ".join(rendered) + " |")
        if row_index == 0:
            rows.append("| " + " | ".join("---" for _ in rendered) + " |")
    result = "\n".join(rows)
    return result[:MARKDOWN_LIMIT], len(result) > MARKDOWN_LIMIT


def _spreadsheet_read(
    invocation: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    arguments = invocation["arguments"]
    source = _materialized_input(invocation, str(arguments["input_id"]))
    values_book = _open_workbook(source, data_only=True, read_only=True)
    formulas_book = _open_workbook(source, data_only=False, read_only=True)
    try:
        values_sheet = _sheet(values_book, arguments.get("sheet"))
        formulas_sheet = _sheet(formulas_book, arguments.get("sheet"))
        requested = str(
            arguments.get("range")
            or f"A1:{get_column_letter(max(1, formulas_sheet.max_column))}{max(1, formulas_sheet.max_row)}"
        ).upper()
        max_cells = max(1, min(MAX_CELLS, int(arguments.get("max_cells", 1000))))
        min_row, min_col, max_row, max_col, actual, range_truncated = _bounded_range(
            requested, max_cells
        )
        values: list[list[Any]] = []
        formulas: list[list[str | None]] = []
        types: list[list[str]] = []
        formats: list[list[str]] = []
        missing_formula_cache = False
        for row in range(min_row, max_row + 1):
            value_row: list[Any] = []
            formula_row: list[str | None] = []
            type_row: list[str] = []
            format_row: list[str] = []
            for column in range(min_col, max_col + 1):
                formula_cell = formulas_sheet.cell(row=row, column=column)
                cached_cell = values_sheet.cell(row=row, column=column)
                formula = (
                    formula_cell.value
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("=")
                    else None
                )
                cached = cached_cell.value
                missing_formula_cache |= formula is not None and cached is None
                value_row.append(_json_value(cached if formula is not None else formula_cell.value))
                formula_row.append(formula)
                type_row.append(
                    _value_type(
                        cached if formula is not None else formula_cell.value,
                        formula=formula is not None,
                        error=formula_cell.data_type == "e",
                    )
                )
                format_row.append(str(formula_cell.number_format))
            values.append(value_row)
            formulas.append(formula_row)
            types.append(type_row)
            formats.append(format_row)
        markdown, markdown_truncated = _markdown(values, formulas)
        warnings = ["formula cache is missing; render to recalculate"] if missing_formula_cache else []
        return (
            {
                "sheets": [
                    {
                        "name": sheet.title,
                        "max_row": int(sheet.max_row),
                        "max_column": int(sheet.max_column),
                    }
                    for sheet in formulas_book.worksheets
                ],
                "sheet": formulas_sheet.title,
                "requested_range": requested,
                "range": actual,
                "values": values,
                "formulas": formulas,
                "types": types,
                "number_formats": formats,
                "markdown": markdown,
                "truncated": range_truncated or markdown_truncated,
                "timezone_policy": "UTC",
                "numeric_policy": "IEEE-754 binary64",
                "warnings": warnings,
            },
            [],
        )
    finally:
        values_book.close()
        formulas_book.close()


def _color(value: str) -> str:
    color = value.strip().removeprefix("#").upper()
    if len(color) not in {6, 8} or any(character not in "0123456789ABCDEF" for character in color):
        raise SpreadsheetActionError("invalid_invocation", "style color must be RRGGBB or AARRGGBB")
    return color


def _apply_style(target, style: dict[str, Any]) -> None:
    if any(key in style for key in ("bold", "italic", "font_size", "font_color")):
        target.font = Font(
            name=target.font.name,
            size=style.get("font_size", target.font.sz),
            bold=style.get("bold", target.font.bold),
            italic=style.get("italic", target.font.italic),
            color=(
                _color(str(style["font_color"]))
                if "font_color" in style
                else target.font.color
            ),
        )
    if "fill_color" in style:
        target.fill = PatternFill("solid", fgColor=_color(str(style["fill_color"])))
    if "horizontal" in style:
        alignment = str(style["horizontal"])
        if alignment not in {"left", "center", "right"}:
            raise SpreadsheetActionError("invalid_invocation", "horizontal alignment is invalid")
        target.alignment = Alignment(horizontal=alignment, vertical=target.alignment.vertical)
    if "number_format" in style:
        target.number_format = str(style["number_format"])


def _decoded_cell(item: dict[str, Any]) -> Any:
    if "formula" in item:
        formula = str(item["formula"])
        if not formula.startswith("="):
            raise SpreadsheetActionError("invalid_invocation", "formula must start with =")
        return formula
    value = item.get("value")
    value_type = item.get("value_type")
    try:
        if value_type == "date":
            return date.fromisoformat(str(value))
        if value_type == "datetime":
            parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            if parsed.tzinfo is None:
                parsed = parsed.replace(tzinfo=timezone.utc)
            return parsed.astimezone(timezone.utc).replace(tzinfo=None)
    except ValueError as exc:
        raise SpreadsheetActionError("invalid_invocation", "date or datetime value is invalid") from exc
    return value


def _set_cell(sheet, item: dict[str, Any]) -> None:
    target = sheet[_cell(str(item["cell"]))]
    target.value = _decoded_cell(item)
    if "number_format" in item:
        target.number_format = str(item["number_format"])
    if item.get("style"):
        _apply_style(target, dict(item["style"]))


def _chart(sheet, specification: dict[str, Any]) -> None:
    chart_type = specification["type"]
    chart = {"bar": BarChart, "line": LineChart, "pie": PieChart}.get(chart_type)
    if chart is None:
        raise SpreadsheetActionError("invalid_invocation", "chart type is unsupported")
    result = chart()
    min_row, min_col, max_row, max_col = _limited_range(str(specification["data_range"]))
    result.add_data(
        Reference(sheet, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row),
        titles_from_data=bool(specification.get("titles_from_data", True)),
    )
    if specification.get("categories_range"):
        cat_min_row, cat_min_col, cat_max_row, cat_max_col = _limited_range(
            str(specification["categories_range"])
        )
        result.set_categories(
            Reference(
                sheet,
                min_col=cat_min_col,
                min_row=cat_min_row,
                max_col=cat_max_col,
                max_row=cat_max_row,
            )
        )
    result.title = str(specification.get("title") or "")
    sheet.add_chart(result, _cell(str(specification.get("anchor", "E2"))))


def _prepare_workbook(workbook) -> None:
    fixed = datetime(1980, 1, 1)
    workbook.properties.created = fixed
    workbook.properties.modified = fixed
    workbook.properties.lastModifiedBy = "SheJane"
    workbook.calculation.calcMode = "auto"
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _normalize_archive(path: Path) -> None:
    normalized = path.with_suffix(".normalized.xlsx")
    try:
        with (
            zipfile.ZipFile(path, "r") as source,
            zipfile.ZipFile(normalized, "w", zipfile.ZIP_DEFLATED, compresslevel=9) as target,
        ):
            for original in sorted(source.infolist(), key=lambda item: item.filename):
                if original.is_dir():
                    continue
                data = source.read(original)
                if original.filename == "docProps/core.xml":
                    data = re.sub(
                        rb"(<dcterms:(?:created|modified)\b[^>]*>)[^<]*(</dcterms:(?:created|modified)>)",
                        rb"\g<1>1980-01-01T00:00:00Z\g<2>",
                        data,
                    )
                info = zipfile.ZipInfo(original.filename, date_time=(1980, 1, 1, 0, 0, 0))
                info.compress_type = zipfile.ZIP_DEFLATED
                info.create_system = 3
                info.external_attr = 0o600 << 16
                target.writestr(info, data)
        os.replace(normalized, path)
    except (OSError, zipfile.BadZipFile) as exc:
        raise SpreadsheetActionError(
            "spreadsheet_write_failed", "XLSX output could not be normalized"
        ) from exc
    finally:
        normalized.unlink(missing_ok=True)


def _atomic_save(workbook, filename: str) -> Path:
    output = _output_root()
    target = output / _safe_output_name(filename)
    if target.exists():
        raise SpreadsheetActionError("spreadsheet_write_failed", "XLSX output already exists")
    descriptor, temporary_name = tempfile.mkstemp(prefix=".xlsx-", suffix=".xlsx", dir=output)
    os.close(descriptor)
    temporary = Path(temporary_name)
    try:
        _prepare_workbook(workbook)
        workbook.save(temporary)
        _normalize_archive(temporary)
        verified = _open_workbook(temporary, read_only=True)
        verified.close()
        os.replace(temporary, target)
        return target
    except SpreadsheetActionError:
        raise
    except Exception as exc:
        raise SpreadsheetActionError(
            "spreadsheet_write_failed", "XLSX output could not be written"
        ) from exc
    finally:
        temporary.unlink(missing_ok=True)


def _spreadsheet_create(
    invocation: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    arguments = invocation["arguments"]
    workbook = Workbook()
    workbook.remove(workbook.active)
    cell_count = 0
    chart_count = 0
    try:
        worksheets = arguments.get("worksheets")
        if not isinstance(worksheets, list) or not 1 <= len(worksheets) <= 64:
            raise SpreadsheetActionError("invalid_invocation", "one to 64 worksheets are required")
        for specification in worksheets:
            try:
                sheet = workbook.create_sheet(str(specification["name"]))
            except (ValueError, KeyError) as exc:
                raise SpreadsheetActionError(
                    "invalid_invocation", "worksheet name is invalid or duplicated"
                ) from exc
            for item in specification.get("cells", []):
                if cell_count >= MAX_CELLS:
                    raise SpreadsheetActionError(
                        "invalid_invocation", "workbook exceeds the action cell limit"
                    )
                _set_cell(sheet, item)
                cell_count += 1
            for merged_range in specification.get("merges", []):
                _limited_range(str(merged_range))
                sheet.merge_cells(str(merged_range).upper())
            for column, width in specification.get("column_widths", {}).items():
                numeric_width = float(width)
                if not 1 <= numeric_width <= 255:
                    raise SpreadsheetActionError("invalid_invocation", "column width is invalid")
                sheet.column_dimensions[_column(str(column))].width = numeric_width
            for chart in specification.get("charts", []):
                _chart(sheet, chart)
                chart_count += 1
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
        target = _atomic_save(workbook, str(arguments["filename"]))
        return (
            {
                "filename": target.name,
                "sheet_count": len(workbook.sheetnames),
                "cell_count": cell_count,
                "chart_count": chart_count,
                "timezone_policy": "UTC",
                "numeric_policy": "IEEE-754 binary64",
                "warnings": [],
            },
            [{"path": f"/output/{target.name}", "media_type": XLSX_MIME, "name": target.name}],
        )
    finally:
        workbook.close()


def _format_range(sheet, range_name: str, style: dict[str, Any]) -> int:
    min_row, min_col, max_row, max_col = _limited_range(range_name)
    count = 0
    for row in sheet.iter_rows(
        min_row=min_row, min_col=min_col, max_row=max_row, max_col=max_col
    ):
        for target in row:
            _apply_style(target, style)
            count += 1
    return count


def _spreadsheet_edit(
    invocation: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    arguments = invocation["arguments"]
    source = _materialized_input(invocation, str(arguments["input_id"]))
    workbook = _open_workbook(source)
    changes: list[str] = []
    try:
        operations = arguments.get("operations")
        if not isinstance(operations, list) or not 1 <= len(operations) <= 128:
            raise SpreadsheetActionError("invalid_invocation", "one to 128 operations are required")
        for operation in operations:
            operation_type = operation["type"]
            sheet = _sheet(workbook, operation.get("sheet"))
            if operation_type == "set_cells":
                min_row, min_col, max_row, max_col = _range(str(operation["range"]))
                rows = operation["values"]
                if sum(len(row) for row in rows) > MAX_CELLS:
                    raise SpreadsheetActionError(
                        "invalid_invocation", "set_cells exceeds the action cell limit"
                    )
                if len(rows) > max_row - min_row + 1 or any(
                    len(row) > max_col - min_col + 1 for row in rows
                ):
                    raise SpreadsheetActionError(
                        "invalid_invocation", "set_cells values exceed the target range"
                    )
                count = 0
                for row_offset, row in enumerate(rows):
                    for column_offset, value in enumerate(row):
                        sheet.cell(min_row + row_offset, min_col + column_offset).value = value
                        count += 1
                changes.append(f"set_cells:{count}")
            elif operation_type == "set_formula":
                formula = str(operation["formula"])
                if not formula.startswith("="):
                    raise SpreadsheetActionError("invalid_invocation", "formula must start with =")
                sheet[_cell(str(operation["cell"]))] = formula
                changes.append("set_formula:1")
            elif operation_type == "format_range":
                count = _format_range(sheet, str(operation["range"]), dict(operation["style"]))
                changes.append(f"format_range:{count}")
            elif operation_type in {"merge_cells", "unmerge_cells"}:
                target_range = str(operation["range"]).upper()
                _limited_range(target_range)
                if operation_type == "merge_cells":
                    sheet.merge_cells(target_range)
                else:
                    sheet.unmerge_cells(target_range)
                changes.append(f"{operation_type}:1")
            elif operation_type == "append_rows":
                if sum(len(row) for row in operation["rows"]) > MAX_CELLS:
                    raise SpreadsheetActionError(
                        "invalid_invocation", "append_rows exceeds the action cell limit"
                    )
                for row in operation["rows"]:
                    sheet.append(row)
                changes.append(f"append_rows:{len(operation['rows'])}")
            elif operation_type == "insert_rows":
                index = int(operation["row"])
                rows = operation.get("rows", [])
                if sum(len(row) for row in rows) > MAX_CELLS:
                    raise SpreadsheetActionError(
                        "invalid_invocation", "insert_rows exceeds the action cell limit"
                    )
                amount = max(1, int(operation.get("amount", len(rows) or 1)))
                sheet.insert_rows(index, amount)
                for row_offset, row in enumerate(rows[:amount]):
                    for column, value in enumerate(row, start=1):
                        sheet.cell(index + row_offset, column).value = value
                changes.append(f"insert_rows:{amount}")
            elif operation_type == "add_chart":
                _chart(sheet, operation["chart"])
                changes.append("add_chart:1")
            else:
                raise SpreadsheetActionError(
                    "invalid_invocation", "spreadsheet edit operation is unsupported"
                )
        default_name = f"{source.stem}.edited.xlsx"
        target = _atomic_save(
            workbook, str(arguments.get("output_filename") or default_name)
        )
        return (
            {
                "filename": target.name,
                "operation_count": len(operations),
                "changes": changes,
                "warnings": [],
            },
            [{"path": f"/output/{target.name}", "media_type": XLSX_MIME, "name": target.name}],
        )
    finally:
        workbook.close()


def _runtime_tool_paths() -> tuple[Path, Path]:
    try:
        mapping = json.loads(os.environ["SHEJANE_PLUGIN_RUNTIME_ASSETS"])
        root = Path(mapping[RUNTIME_ASSET_ID]).resolve(strict=True)
        config = json.loads((root / "office-runtime.json").read_text(encoding="utf-8"))
    except (KeyError, OSError, ValueError, json.JSONDecodeError) as exc:
        raise SpreadsheetActionError(
            "engine_unavailable", "Office runtime asset is unavailable"
        ) from exc
    tools = []
    for key in ("soffice", "mutool"):
        try:
            relative = PurePosixPath(str(config[key]))
            if relative.is_absolute() or any(part in {"", ".", ".."} for part in relative.parts):
                raise ValueError
            candidate = root.joinpath(*relative.parts).resolve(strict=True)
            candidate.relative_to(root)
        except (KeyError, OSError, ValueError) as exc:
            raise SpreadsheetActionError(
                "engine_unavailable", "Office runtime asset is invalid"
            ) from exc
        if not candidate.is_file() or not os.access(candidate, os.X_OK):
            raise SpreadsheetActionError(
                "engine_unavailable", "Office runtime executable is unavailable"
            )
        tools.append(candidate)
    return tools[0], tools[1]


def _run_tool(command: list[str], *, cwd: Path, deadline: float, env: dict[str, str]) -> str:
    remaining = deadline - time.monotonic()
    if remaining <= 0:
        raise SpreadsheetActionError("render_failed", "spreadsheet renderer timed out")
    try:
        completed = subprocess.run(
            command,
            cwd=cwd,
            env=env,
            stdin=subprocess.DEVNULL,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            timeout=remaining,
            check=False,
        )
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise SpreadsheetActionError(
            "render_failed", "spreadsheet renderer did not complete"
        ) from exc
    if completed.returncode != 0:
        raise SpreadsheetActionError("render_failed", "spreadsheet renderer rejected the input")
    return completed.stdout[-16_384:].decode("utf-8", errors="replace")


def _office_command(soffice: Path, profile: Path) -> list[str]:
    return [
        str(soffice),
        "--headless",
        "--nologo",
        "--nodefault",
        "--nolockcheck",
        "--norestore",
        f"-env:UserInstallation={profile.as_uri()}",
    ]


def _spreadsheet_render(
    invocation: dict[str, Any],
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    arguments = invocation["arguments"]
    source = _materialized_input(invocation, str(arguments["input_id"]))
    verified = _open_workbook(source, read_only=True)
    verified.close()
    soffice, mutool = _runtime_tool_paths()
    output = _output_root()
    profile = output / ".libreoffice-profile"
    temporary = output / ".runtime-tmp"
    work = output / ".work"
    for directory in (profile, temporary, work):
        directory.mkdir(mode=0o700, exist_ok=False)
    deadline = time.monotonic() + max(
        1.0, min(300.0, float(invocation["limits"]["timeout_ms"]) / 1000 * 0.8)
    )
    env = {
        "HOME": str(profile),
        "TMPDIR": str(temporary),
        "TEMP": str(temporary),
        "TMP": str(temporary),
        "LANG": "C.UTF-8",
        "LC_ALL": "C.UTF-8",
        "TZ": "UTC",
        "PATH": os.pathsep.join(sorted({str(soffice.parent), str(mutool.parent)})),
    }
    try:
        recalculated_name = _safe_output_name(f"{source.stem}.recalculated.xlsx")
        working_copy = work / recalculated_name
        shutil.copyfile(source, working_copy)
        _run_tool(
            [
                *_office_command(soffice, profile),
                "--convert-to",
                "xlsx:Calc MS Excel 2007 XML",
                "--outdir",
                str(output),
                str(working_copy),
            ],
            cwd=output,
            deadline=deadline,
            env=env,
        )
        recalculated = output / recalculated_name
        recalculated_book = _open_workbook(recalculated, data_only=True, read_only=True)
        recalculated_book.close()
        _normalize_archive(recalculated)
        _run_tool(
            [
                *_office_command(soffice, profile),
                "--convert-to",
                "pdf:calc_pdf_Export",
                "--outdir",
                str(output),
                str(recalculated),
            ],
            cwd=output,
            deadline=deadline,
            env=env,
        )
        pdf = output / f"{recalculated.stem}.pdf"
        if not pdf.is_file() or not pdf.read_bytes().startswith(b"%PDF"):
            raise SpreadsheetActionError(
                "render_failed", "spreadsheet renderer produced no valid PDF"
            )
        info = _run_tool(
            [str(mutool), "info", str(pdf)], cwd=output, deadline=deadline, env=env
        )
        page_match = re.search(r"(?:Pages|pages)\s*:\s*(\d+)", info)
        if page_match is None or int(page_match.group(1)) < 1:
            raise SpreadsheetActionError("render_failed", "PDF page count is unavailable")
        page_count = int(page_match.group(1))
        include_png = bool(arguments.get("include_png", True))
        rendered_pages = min(page_count, int(arguments.get("max_pages", 20))) if include_png else 0
        png_names: list[str] = []
        if rendered_pages:
            pattern = output / f"{recalculated.stem}.page-%04d.png"
            _run_tool(
                [
                    str(mutool),
                    "draw",
                    "-q",
                    "-r",
                    str(int(arguments.get("dpi", 144))),
                    "-o",
                    str(pattern),
                    str(pdf),
                    f"1-{rendered_pages}",
                ],
                cwd=output,
                deadline=deadline,
                env=env,
            )
            for page in range(1, rendered_pages + 1):
                name = f"{recalculated.stem}.page-{page:04d}.png"
                if not (output / name).is_file():
                    raise SpreadsheetActionError(
                        "render_failed", "PNG page rendering was incomplete"
                    )
                png_names.append(name)
        warnings = (
            ["PNG preview limited by max_pages"]
            if include_png and rendered_pages < page_count
            else []
        )
        artifacts = [
            {
                "path": f"/output/{recalculated.name}",
                "media_type": XLSX_MIME,
                "name": recalculated.name,
            },
            {"path": f"/output/{pdf.name}", "media_type": "application/pdf", "name": pdf.name},
            *(
                {"path": f"/output/{name}", "media_type": "image/png", "name": name}
                for name in png_names
            ),
        ]
        return (
            {
                "page_count": page_count,
                "rendered_pages": rendered_pages,
                "recalculated_name": recalculated.name,
                "pdf_name": pdf.name,
                "png_names": png_names,
                "warnings": warnings,
            },
            artifacts,
        )
    finally:
        for directory in (profile, temporary, work):
            shutil.rmtree(directory, ignore_errors=True)


def _success(
    invocation: dict[str, Any],
    output: dict[str, Any],
    artifacts: list[dict[str, str]],
) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "invocation_id": invocation["invocation_id"],
        "operation_id": invocation["operation_id"],
        "status": "succeeded",
        "output": output,
        "artifacts": artifacts,
    }


def _failure(invocation: dict[str, Any], error: SpreadsheetActionError) -> dict[str, Any]:
    return {
        "schema_version": 1,
        "invocation_id": invocation["invocation_id"],
        "operation_id": invocation["operation_id"],
        "status": "failed",
        "error": {
            "code": error.code,
            "message": str(error),
            "retryable": error.retryable,
        },
        "artifacts": [],
    }


def _invoke(invocation: dict[str, Any]) -> dict[str, Any]:
    action = invocation.get("action")
    if not isinstance(action, dict) or action.get("plugin_id") != PLUGIN_ID:
        raise SpreadsheetActionError("invalid_invocation", "plugin identity does not match")
    handlers = {
        "spreadsheet.read": _spreadsheet_read,
        "spreadsheet.create": _spreadsheet_create,
        "spreadsheet.edit": _spreadsheet_edit,
        "spreadsheet.render": _spreadsheet_render,
    }
    handler = handlers.get(str(action.get("action_id") or ""))
    if handler is None:
        raise SpreadsheetActionError("invalid_invocation", "spreadsheet Action is unsupported")
    output, artifacts = handler(invocation)
    return _success(invocation, output, artifacts)


def main() -> None:
    initialize = _request("initialize")
    params = initialize["params"]
    if (
        params.get("protocol_version") != 1
        or params.get("plugin_id") != PLUGIN_ID
        or not set(params.get("actions", ())).issubset(ACTION_IDS)
    ):
        raise SpreadsheetActionError("incompatible_runtime", "worker initialization is incompatible")
    _reply(
        int(initialize["id"]),
        {
            "protocol_version": 1,
            "process_isolated": True,
            "access_isolated": os.environ.get("SHEJANE_PLUGIN_ACCESS_ISOLATED") == "1",
            "resource_isolated": os.environ.get("SHEJANE_PLUGIN_RESOURCE_ISOLATED") == "1",
            "sandboxed": os.environ.get("SHEJANE_PLUGIN_SANDBOXED") == "1",
        },
    )
    invoke = _request("invoke")
    invocation = invoke["params"]
    try:
        result = _invoke(invocation)
    except SpreadsheetActionError as exc:
        result = _failure(invocation, exc)
    except Exception as exc:
        if os.environ.get("SHEJANE_PLUGIN_DEBUG") == "1":
            print(f"{type(exc).__name__}: {exc}", file=sys.stderr, flush=True)
        result = _failure(
            invocation,
            SpreadsheetActionError("plugin_failed", "spreadsheet worker failed"),
        )
    _reply(int(invoke["id"]), result)
    shutdown = _request("shutdown")
    _reply(int(shutdown["id"]), {})


if __name__ == "__main__":
    main()
